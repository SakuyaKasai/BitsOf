# Excel版/process_batch_base.py - プロセス別バッチ処理のベースクラス
"""
プロセス別exe用のベースクラス
各プロセス用のbatch.pyはこれを継承して作成
"""

import pandas as pd
import sys
import os

# 共通基盤へのパスを追加
sys.path.append(os.path.join(os.path.dirname(__file__), '..', '共通基盤'))
import backend

import subprocess
import time
from datetime import datetime
from abc import ABC, abstractmethod

class ProcessBatchBase(ABC):
    """プロセス別バッチ処理のベースクラス"""
    
    def __init__(self):
        self.is_exe = getattr(sys, 'frozen', False)
    
    @property
    @abstractmethod
    def process_type(self):
        """処理タイプ名を返す（サブクラスで実装）"""
        pass
    
    @property
    @abstractmethod
    def process_description(self):
        """処理の説明を返す（サブクラスで実装）"""
        pass
    
    @property
    @abstractmethod
    def usage_instructions(self):
        """使用方法を返す（サブクラスで実装）"""
        pass
    
    @property
    @abstractmethod
    def input_examples(self):
        """入力例のリストを返す（サブクラスで実装）"""
        pass
    
    def create_template_excel(self, filepath):
        """
        プロセス専用Excelテンプレートを作成
        """
        try:
            # プロセス情報取得
            process_info = backend.get_process_info(self.process_type)
            
            # テンプレートデータ作成
            template_data = []
            
            # ヘッダー部分
            template_data.extend([
                [self.process_type, '', '', '', '', '', ''],  # A1: システム名
                [self.process_description, '', '', '', '', '', ''],  # A2: 処理説明
                [self.usage_instructions, '', '', '', '', '', ''],  # A3: 使用方法
                ['', '', '', '', '', '', ''],  # A4: 空行
                ['【入力データ】', '番号', '完了日時', '結果', '詳細', '備考', ''],  # A5: ヘッダー行
            ])
            
            # 入力例の追加
            for i, example in enumerate(self.input_examples[:5], 1):
                template_data.append([example, '', '', '', '', '', ''])
            
            # 空行の追加（入力用）
            for i in range(10):
                template_data.append(['', '', '', '', '', '', ''])
            
            # DataFrameに変換
            columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
            df = pd.DataFrame(template_data, columns=columns)
            
            # Excelファイルに保存（書式設定付き）
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='処理', index=False, header=False)
                
                # ワークシートの取得と書式設定
                worksheet = writer.sheets['処理']
                
                # 列幅調整
                worksheet.column_dimensions['A'].width = 30  # 入力データ
                worksheet.column_dimensions['B'].width = 8   # 番号
                worksheet.column_dimensions['C'].width = 18  # 完了日時
                worksheet.column_dimensions['D'].width = 10  # 結果
                worksheet.column_dimensions['E'].width = 40  # 詳細
                worksheet.column_dimensions['F'].width = 15  # 備考
                
                # セルの書式設定
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                # スタイル定義
                title_font = Font(size=14, bold=True, color="FFFFFF")
                title_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
                
                header_font = Font(size=11, bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")
                
                desc_font = Font(size=10, color="333333")
                desc_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
                
                center_alignment = Alignment(horizontal="center", vertical="center")
                left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # A1: システム名
                cell_a1 = worksheet['A1']
                cell_a1.font = title_font
                cell_a1.fill = title_fill
                cell_a1.alignment = center_alignment
                cell_a1.border = thin_border
                worksheet.merge_cells('A1:G1')
                
                # A2-A3: 説明部分
                for row in range(2, 4):
                    cell = worksheet[f'A{row}']
                    cell.font = desc_font
                    cell.fill = desc_fill
                    cell.alignment = left_alignment
                    cell.border = thin_border
                    worksheet.merge_cells(f'A{row}:G{row}')
                
                # A5: ヘッダー行
                for col_idx, col_letter in enumerate(['A', 'B', 'C', 'D', 'E', 'F'], 1):
                    cell = worksheet[f'{col_letter}5']
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    cell.border = thin_border
                
                # 入力エリアの枠線
                for row in range(6, len(template_data) + 1):
                    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
                        cell = worksheet[f'{col_letter}{row}']
                        cell.border = thin_border
                        if col_letter == 'A':
                            cell.alignment = left_alignment
                        else:
                            cell.alignment = center_alignment
            
            print(f"📝 {self.process_type}用テンプレートファイル '{filepath}' を作成しました")
            return True
            
        except Exception as e:
            print(f"❌ テンプレート作成エラー: {e}")
            return False
    
    def open_excel_for_editing(self, filepath):
        """Excelファイルを開いてユーザーに編集させる"""
        print("=" * 60)
        print(f"📊 {self.process_type} - Excel入力モード")
        print("=" * 60)
        print(f"📂 {filepath} を開きます")
        print("💡 A列（入力データ）に処理したい内容を記入してください")
        print("💡 A6行目以降に入力してください")
        print("💡 編集完了後、Excelを保存して閉じてください")
        print("💡 その後、このプログラムでEnterキーを押してください")
        
        try:
            # OS判定してExcelを開く
            if sys.platform.startswith('win'):
                os.startfile(filepath)
            elif sys.platform.startswith('darwin'):  # macOS
                subprocess.run(['open', filepath])
            else:  # Linux
                subprocess.run(['xdg-open', filepath])
                
        except Exception as e:
            print(f"⚠️ Excelの自動起動に失敗: {e}")
            print(f"手動で {filepath} を開いて編集してください")
        
        # ユーザーの編集完了待ち
        input("\n⏸️ 編集完了後、Enterキーを押してください...")
    
    def load_and_validate_excel(self, filepath):
        """編集されたExcelファイルを読み込み・検証"""
        try:
            # Excelファイル読み込み（A列のみ、6行目以降）
            df = pd.read_excel(filepath, header=None, usecols=[0], skiprows=5)
            df.columns = ['入力データ']
            
            # 空行や無効な行を除去
            df = df.dropna(subset=['入力データ'])
            df = df[df['入力データ'].astype(str).str.strip() != '']
            
            if len(df) == 0:
                print("❌ 有効な入力データがありません")
                return None
            
            print(f"📊 {len(df)} 行の有効なデータを読み込みました")
            
            # データの表示
            print(f"\n📋 読み込まれたデータ ({self.process_type}):")
            for idx, row in df.iterrows():
                data_preview = str(row['入力データ'])[:50]
                if len(str(row['入力データ'])) > 50:
                    data_preview += "..."
                print(f"  {idx+1}. {data_preview}")
            
            return df
            
        except Exception as e:
            print(f"❌ Excelファイル読み込みエラー: {e}")
            return None
    
    def execute_processing(self, input_df):
        """処理実行"""
        print("\n" + "=" * 60)
        print(f"🔄 {self.process_type} 処理開始")
        print("=" * 60)
        
        # 入力データを結合
        input_data = "\n".join(input_df['入力データ'].astype(str).tolist())
        
        # 入力検証
        is_valid, error_msg = backend.validate_input(self.process_type, input_data)
        if not is_valid:
            print(f"❌ 入力検証エラー: {error_msg}")
            return None
        
        # 処理実行
        try:
            start_time = datetime.now()
            print(f"⏰ 処理開始: {start_time.strftime('%H:%M:%S')}")
            
            results = backend.execute_process(self.process_type, input_data)
            
            end_time = datetime.now()
            duration = (end_time - start_time).total_seconds()
            
            success_count = len([r for r in results if r[2] == "成功"])
            failure_count = len([r for r in results if r[2] == "失敗"])
            
            print(f"✅ 処理完了: 成功{success_count}件, 失敗{failure_count}件 ({duration:.1f}秒)")
            
            return results
            
        except Exception as e:
            print(f"❌ 処理エラー: {e}")
            return None
    
    def save_results_to_excel(self, input_df, results, filepath):
        """結果をExcelファイルに保存"""
        try:
            # 元のテンプレートを読み込み
            workbook = pd.read_excel(filepath, sheet_name='処理', header=None)
            
            # 結果を書き込み（B列から）
            with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                # 結果をセルに直接書き込み
                worksheet = writer.book['処理']
                
                # 結果の書き込み（6行目から）
                for i, result in enumerate(results):
                    row_num = i + 6  # 6行目から開始
                    
                    # B列: 番号
                    worksheet[f'B{row_num}'] = result[0]
                    # C列: 完了日時
                    worksheet[f'C{row_num}'] = result[1]
                    # D列: 結果
                    worksheet[f'D{row_num}'] = result[2]
                    # E列: 詳細
                    worksheet[f'E{row_num}'] = result[3]
                    # F列: 備考
                    worksheet[f'F{row_num}'] = result[4]
                
                # 統計情報を下部に追加
                stats_start_row = len(results) + 8
                worksheet[f'A{stats_start_row}'] = "【処理統計】"
                worksheet[f'A{stats_start_row + 1}'] = f"総件数: {len(results)}"
                worksheet[f'A{stats_start_row + 2}'] = f"成功: {len([r for r in results if r[2] == '成功'])}"
                worksheet[f'A{stats_start_row + 3}'] = f"失敗: {len([r for r in results if r[2] == '失敗'])}"
                
                # 統計部分の書式設定
                from openpyxl.styles import Font, PatternFill
                stats_font = Font(bold=True, color="333333")
                stats_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                
                for row in range(stats_start_row, stats_start_row + 4):
                    cell = worksheet[f'A{row}']
                    cell.font = stats_font
                    cell.fill = stats_fill
            
            print(f"✅ 結果を {filepath} に保存しました")
            return True
            
        except Exception as e:
            print(f"❌ ファイル保存エラー: {e}")
            return False
    
    def run(self):
        """メイン実行プロセス"""
        print("=" * 80)
        print(f"🚀 {self.process_type} 処理システム")
        print("=" * 80)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        work_file = f"{self.process_type.replace('処理', '')}_work_{timestamp}.xlsx"
        
        # 1. テンプレート作成
        if not self.create_template_excel(work_file):
            return
        
        # 2. ユーザーに編集させる
        self.open_excel_for_editing(work_file)
        
        # 3. 編集されたファイルを読み込み
        input_df = self.load_and_validate_excel(work_file)
        if input_df is None:
            return
        
        # 4. 処理実行
        results = self.execute_processing(input_df)
        if results is None:
            return
        
        # 5. 結果保存
        if self.save_results_to_excel(input_df, results, work_file):
            # 結果ファイルを開く
            try:
                print(f"\n💡 結果ファイル {work_file} を開きますか？ (y/N): ", end="")
                if input().strip().lower() == 'y':
                    if sys.platform.startswith('win'):
                        os.startfile(work_file)
                    elif sys.platform.startswith('darwin'):
                        subprocess.run(['open', work_file])
                    else:
                        subprocess.run(['xdg-open', work_file])
            except:
                pass
        
        print(f"\n🎉 {self.process_type} 処理完了")