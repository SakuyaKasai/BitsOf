import requests
import pandas as pd
from datetime import datetime
import os
import sys
import platform
import threading
import time
from urllib.parse import urlparse

class URLProcessor:
    def __init__(self):
        self.timeout = 10  # リクエストタイムアウト（秒）
        
    def process_url(self, url):
        """
        URLを処理して情報を取得
        Windows対応版 - signal.alarmを使用しない
        """
        result = {
            'url': url,
            'status_code': None,
            'title': None,
            'description': None,
            'keywords': None,
            'error': None
        }
        
        try:
            # URLの形式チェック
            if not url.startswith(('http://', 'https://')):
                url = 'https://' + url
                result['url'] = url
            
            # Windowsでも動作するタイムアウト付きリクエスト
            response = requests.get(
                url, 
                timeout=self.timeout,
                headers={
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
                }
            )
            
            result['status_code'] = response.status_code
            
            if response.status_code == 200:
                # HTMLからタイトル抽出
                content = response.text
                
                # タイトル抽出
                if '<title>' in content and '</title>' in content:
                    title_start = content.find('<title>') + 7
                    title_end = content.find('</title>')
                    result['title'] = content[title_start:title_end].strip()
                
                # メタディスクリプション抽出
                if 'name="description"' in content:
                    desc_start = content.find('name="description"')
                    desc_section = content[desc_start:desc_start+500]
                    if 'content="' in desc_section:
                        content_start = desc_section.find('content="') + 9
                        content_end = desc_section.find('"', content_start)
                        result['description'] = desc_section[content_start:content_end].strip()
                
                # キーワード抽出
                if 'name="keywords"' in content:
                    kw_start = content.find('name="keywords"')
                    kw_section = content[kw_start:kw_start+500]
                    if 'content="' in kw_section:
                        content_start = kw_section.find('content="') + 9
                        content_end = kw_section.find('"', content_start)
                        result['keywords'] = kw_section[content_start:content_end].strip()
            
        except requests.exceptions.Timeout:
            result['error'] = f'タイムアウト (>{self.timeout}秒)'
        except requests.exceptions.ConnectionError:
            result['error'] = '接続エラー'
        except requests.exceptions.RequestException as e:
            result['error'] = f'リクエストエラー: {str(e)}'
        except Exception as e:
            result['error'] = f'予期しないエラー: {str(e)}'
        
        return result

def create_excel_template():
    """Excelテンプレートファイルを作成"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"URL処理_{timestamp}.xlsx"
    
    # サンプルデータ
    sample_data = {
        'A': ['', '', '📝 URL処理システム', '', '⬇️ A6行目以降にURLを入力', 
              'https://example.com', 'https://google.com', 'https://github.com', 
              'https://stackoverflow.com', 'https://python.org']
    }
    
    df = pd.DataFrame(sample_data)
    
    # Excelファイル作成
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='入力', index=False, header=False)
        
        # ワークシートの調整
        workbook = writer.book
        worksheet = writer.sheets['入力']
        
        # 列幅調整
        worksheet.column_dimensions['A'].width = 50
        
        # セルスタイル適用
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # ヘッダー行のスタイル
        for row in range(1, 6):
            cell = worksheet[f'A{row}']
            if row == 3:  # タイトル行
                cell.font = Font(bold=True, size=14)
                cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            elif row == 5:  # 説明行
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    
    return filename

def read_input_data(filename):
    """入力データをExcelファイルから読み込み"""
    try:
        df = pd.read_excel(filename, sheet_name='入力', header=None)
        
        # A6行目以降のデータを取得（0ベースなので5行目以降）
        urls = []
        for i in range(5, len(df)):
            if pd.notna(df.iloc[i, 0]):
                url = str(df.iloc[i, 0]).strip()
                if url and not url.startswith('⬇️'):
                    urls.append(url)
        
        return urls
    except Exception as e:
        print(f"❌ ファイル読み込みエラー: {e}")
        return []

def save_results(results, input_filename):
    """処理結果をExcelファイルに保存"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"URL処理結果_{timestamp}.xlsx"
    
    # 結果データフレーム作成
    df_results = pd.DataFrame(results)
    
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        df_results.to_excel(writer, sheet_name='処理結果', index=False)
        
        # ワークシートの調整
        workbook = writer.book
        worksheet = writer.sheets['処理結果']
        
        # 列幅調整
        worksheet.column_dimensions['A'].width = 50  # URL
        worksheet.column_dimensions['B'].width = 15  # ステータス
        worksheet.column_dimensions['C'].width = 50  # タイトル
        worksheet.column_dimensions['D'].width = 80  # 説明
        worksheet.column_dimensions['E'].width = 50  # キーワード
        worksheet.column_dimensions['F'].width = 30  # エラー
        
        # ヘッダー行スタイル
        from openpyxl.styles import Font, PatternFill
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    
    return output_filename

def main():
    print("🚀 URL処理システム")
    print("=" * 60)
    
    # Step 1: Excelテンプレート作成
    print("📝 Step 1: Excelテンプレート作成中...")
    template_file = create_excel_template()
    print(f"📝 テンプレートファイル '{template_file}' を作成しました")
    
    # Step 2: Excel編集モード
    print("📊 Step 2: Excel編集モード...")
    print("=" * 60)
    print("📊 Excel入力モード")
    print("=" * 60)
    print(f"📂 {template_file} を開きます")
    print("💡 A列（入力データ）に処理したいURLを記入してください")
    print("💡 A6行目以降に入力してください")
    print("💡 編集完了後、Excelを保存して閉じてください")
    print("💡 その後、このプログラムでEnterキーを押してください")
    
    # Excelファイルを開く
    if platform.system() == 'Windows':
        os.startfile(template_file)
    elif platform.system() == 'Darwin':  # macOS
        os.system(f'open "{template_file}"')
    else:  # Linux
        os.system(f'xdg-open "{template_file}"')
    
    # ユーザー入力待ち
    input("⏸️ 編集完了後、Enterキーを押してください...")
    
    # Step 3: 入力データ読み込み
    print("📂 Step 3: 入力データ読み込み中...")
    urls = read_input_data(template_file)
    
    if not urls:
        print("❌ 入力データが見つかりませんでした")
        input("Enterキーを押して終了...")
        return
    
    print("📋 読み込まれたデータ:")
    for i, url in enumerate(urls, 1):
        print(f"  {i}. {url}")
    print(f"📊 {len(urls)} 件のURLを読み込みました")
    
    # Step 4: URL処理実行
    print("🔄 Step 4: URL処理実行中...")
    print(f"🔄 {len(urls)} 件のURL処理を開始...")
    
    processor = URLProcessor()
    results = []
    
    for i, url in enumerate(urls, 1):
        print(f"  ({i}/{len(urls)}) 処理中: {url}...")
        result = processor.process_url(url)
        results.append(result)
        
        # 処理状況表示
        if result['error']:
            print(f"    ❌ エラー: {result['error']}")
        else:
            print(f"    ✅ 完了 (ステータス: {result['status_code']})")
    
    # Step 5: 結果保存
    print("💾 Step 5: 結果保存中...")
    output_file = save_results(results, template_file)
    print(f"📁 結果ファイル '{output_file}' を作成しました")
    
    # 結果ファイルを開く
    if platform.system() == 'Windows':
        os.startfile(output_file)
    elif platform.system() == 'Darwin':  # macOS
        os.system(f'open "{output_file}"')
    else:  # Linux
        os.system(f'xdg-open "{output_file}"')
    
    print("✅ 処理完了!")
    print("=" * 60)
    input("Enterキーを押して終了...")

if __name__ == "__main__":
    main()