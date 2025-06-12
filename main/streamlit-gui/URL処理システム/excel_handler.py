# URL処理システム/excel_handler.py - Excel入出力処理
"""
Excelファイルの作成、読み込み、書き込み処理

Excel操作に関する全ての処理をここに集約
"""

import pandas as pd
import os
import sys
import subprocess
from typing import List, Optional

def create_excel_template(filepath: str) -> bool:
    """
    URL処理用のExcelテンプレートを作成
    
    Args:
        filepath (str): 作成するExcelファイルのパス
    
    Returns:
        bool: 作成成功時True
    """
    try:
        # テンプレートデータ作成
        template_data = []
        
        # ヘッダー部分
        template_data.extend([
            ["URL処理システム", '', '', '', '', '', ''],  # A1: システム名
            ["指定されたURLに対して処理を実行し、結果を記録します", '', '', '', '', '', ''],  # A2: 説明
            ["使用方法: A列の6行目以降にURLを入力してください（例: https://example.com）", '', '', '', '', '', ''],  # A3: 使用方法
            ['', '', '', '', '', '', ''],  # A4: 空行
            ['【入力データ】', '番号', '完了日時', '結果', '詳細', '備考', ''],  # A5: ヘッダー行
        ])
        
        # 入力例の追加
        examples = [
            "https://example.com",
            "https://google.com",
            "https://github.com",
            "https://stackoverflow.com",
            "https://python.org"
        ]
        
        for example in examples:
            template_data.append([example, '', '', '', '', '', ''])
        
        # 空行の追加（入力用）
        for i in range(10):
            template_data.append(['', '', '', '', '', '', ''])
        
        # DataFrameに変換
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
        df = pd.DataFrame(template_data, columns=columns)
        
        # Excelファイルに保存（書式設定付き）
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='URL処理', index=False, header=False)
            
            # ワークシートの取得と書式設定
            worksheet = writer.sheets['URL処理']
            
            # 列幅調整
            worksheet.column_dimensions['A'].width = 35  # 入力データ（URL用に広く）
            worksheet.column_dimensions['B'].width = 8   # 番号
            worksheet.column_dimensions['C'].width = 18  # 完了日時
            worksheet.column_dimensions['D'].width = 10  # 結果
            worksheet.column_dimensions['E'].width = 50  # 詳細（URL処理結果用に広く）
            worksheet.column_dimensions['F'].width = 15  # 備考
            
            # セルの書式設定
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # スタイル定義
            title_font = Font(size=14, bold=True, color="FFFFFF")
            title_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
            
            header_font = Font(size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")
            
            desc_font = Font(size=10, color="333333")
            desc_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
            
            center_alignment = Alignment(horizontal="center", vertical="center")
            left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # A1: システム名
            cell_a1 = worksheet['A1']
            cell_a1.font = title_font
            cell_a1.fill = title_fill
            cell_a1.alignment = center_alignment
            cell_a1.border = thin_border
            worksheet.merge_cells('A1:G1')
            
            # A2-A3: 説明部分
            for row in range(2, 4):
                cell = worksheet[f'A{row}']
                cell.font = desc_font
                cell.fill = desc_fill
                cell.alignment = left_alignment
                cell.border = thin_border
                worksheet.merge_cells(f'A{row}:G{row}')
            
            # A5: ヘッダー行
            for col_idx, col_letter in enumerate(['A', 'B', 'C', 'D', 'E', 'F'], 1):
                cell = worksheet[f'{col_letter}5']
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # 入力エリアの枠線
            for row in range(6, len(template_data) + 1):
                for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
                    cell = worksheet[f'{col_letter}{row}']
                    cell.border = thin_border
                    if col_letter == 'A':
                        cell.alignment = left_alignment
                    else:
                        cell.alignment = center_alignment
        
        print(f"📝 テンプレートファイル '{filepath}' を作成しました")
        return True
        
    except Exception as e:
        print(f"❌ テンプレート作成エラー: {e}")
        return False

def open_excel_for_user(filepath: str, data_type: str = "URL"):
    """
    Excelファイルを開いてユーザーに編集してもらう
    
    Args:
        filepath (str): Excelファイルのパス
        data_type (str): データタイプ（表示用）
    """
    print("=" * 60)
    print(f"📊 Excel入力モード")
    print("=" * 60)
    print(f"📂 {filepath} を開きます")
    print(f"💡 A列（入力データ）に処理したい{data_type}を記入してください")
    print("💡 A6行目以降に入力してください")
    print("💡 編集完了後、Excelを保存して閉じてください")
    print("💡 その後、このプログラムでEnterキーを押してください")
    
    try:
        # OS判定してExcelを開く
        if sys.platform.startswith('win'):
            os.startfile(filepath)
        elif sys.platform.startswith('darwin'):  # macOS
            subprocess.run(['open', filepath])
        else:  # Linux
            subprocess.run(['xdg-open', filepath])
            
    except Exception as e:
        print(f"⚠️ Excelの自動起動に失敗: {e}")
        print(f"手動で {filepath} を開いて編集してください")
    
    # ユーザーの編集完了待ち
    input("\n⏸️ 編集完了後、Enterキーを押してください...")

def read_input_data(filepath: str) -> Optional[List[str]]:
    """
    編集されたExcelファイルから入力データを読み込み
    
    Args:
        filepath (str): Excelファイルのパス
    
    Returns:
        Optional[List[str]]: 入力データのリスト（失敗時はNone）
    """
    try:
        # Excelファイル読み込み（A列のみ、6行目以降）
        df = pd.read_excel(filepath, sheet_name='URL処理', header=None, usecols=[0], skiprows=5)
        df.columns = ['入力データ']
        
        # 空行や無効な行を除去
        df = df.dropna(subset=['入力データ'])
        df = df[df['入力データ'].astype(str).str.strip() != '']
        
        if len(df) == 0:
            print("❌ 有効な入力データがありません")
            return None
        
        # データの表示
        input_list = df['入力データ'].astype(str).tolist()
        print(f"\n📋 読み込まれたデータ:")
        for idx, data in enumerate(input_list, 1):
            data_preview = str(data)[:50]
            if len(str(data)) > 50:
                data_preview += "..."
            print(f"  {idx}. {data_preview}")
        
        return input_list
        
    except Exception as e:
        print(f"❌ Excelファイル読み込みエラー: {e}")
        return None

def save_results(filepath: str, results: List[List]) -> bool:
    """
    処理結果をExcelファイルに保存
    
    Args:
        filepath (str): Excelファイルのパス
        results (List[List]): 処理結果のリスト
    
    Returns:
        bool: 保存成功時True
    """
    try:
        # 元のテンプレートを維持しながら結果を書き込み
        with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # ワークシートを取得
            workbook = writer.book
            worksheet = workbook['URL処理']
            
            # 結果の書き込み（6行目から、B列以降）
            for i, result in enumerate(results):
                row_num = i + 6  # 6行目から開始
                
                # B列: 番号
                worksheet[f'B{row_num}'] = result[0]
                # C列: 完了日時
                worksheet[f'C{row_num}'] = result[1]
                # D列: 結果
                worksheet[f'D{row_num}'] = result[2]
                # E列: 詳細
                worksheet[f'E{row_num}'] = result[3]
                # F列: 備考
                worksheet[f'F{row_num}'] = result[4]
            
            # 統計情報を下部に追加
            stats_start_row = len(results) + 8
            worksheet[f'A{stats_start_row}'] = "【処理統計】"
            worksheet[f'A{stats_start_row + 1}'] = f"総件数: {len(results)}"
            
            success_count = len([r for r in results if r[2] == '成功'])
            failure_count = len([r for r in results if r[2] == '失敗'])
            
            worksheet[f'A{stats_start_row + 2}'] = f"成功: {success_count}"
            worksheet[f'A{stats_start_row + 3}'] = f"失敗: {failure_count}"
            
            if len(results) > 0:
                success_rate = (success_count / len(results)) * 100
                worksheet[f'A{stats_start_row + 4}'] = f"成功率: {success_rate:.1f}%"
            
            # 統計部分の書式設定
            from openpyxl.styles import Font, PatternFill
            stats_font = Font(bold=True, color="333333")
            stats_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            
            for row in range(stats_start_row, stats_start_row + 5):
                cell = worksheet[f'A{row}']
                cell.font = stats_font
                cell.fill = stats_fill
        
        print(f"✅ 結果を {filepath} に保存しました")
        return True
        
    except Exception as e:
        print(f"❌ ファイル保存エラー: {e}")
        return False